import re
import shutil
import os
from openpyxl import load_workbook

# script to make folders for each respective participant
# used as destination folder

wb = load_workbook(filename= 'summary_of_participants.xlsx')
ws = wb.active

column_participants =  ws['A']

parent_dir = 'C:\\Users\\justi\\Documents\\certs_landing'

part_array = []

for cell in column_participants:
    part_array.append(cell.value)
    
for item in part_array:
    os.mkdir(os.path.join(parent_dir,item))
    

