import re
import shutil
import os
from openpyxl import load_workbook


# grab each row. with name being first index, and suceeding are all topics
# input this into an array, convert name into string, and all topics as boolean
# loop through array, check if topic is true:
    # if false, move on
    # else, grab index number.
    # match index with file loc
    # move file loc into new directory


# print(directory)

wb = load_workbook(filename= 'summary_of_participants.xlsx')
ws = wb.active


# functions takes in name of participants, and index
def cert_finder(name_of_participant,topic_number):
    main_dir = 'C:\\Users\\justi\\Documents\\certs'

    result = False
    folder_dest = 'C:\\Users\\justi\\Documents\\certs_landing'
    
    # if topic_number in [i for i in range(1,6)]:
        # day one 
        
    dir_of_cert = main_dir + f'\\Topic {str(topic_number)}' 
    # sets the directory of the topic

    print(dir_of_cert)
    
    certs = os.listdir(dir_of_cert)
    # lists the certs inside the topic dir
    for certificate in certs:
        # loops through all certificates inside respective topic
        new_name_certificate = certificate[:-4]
        # print(certificate[:-4])
        if new_name_certificate == name_of_participant:
            # copy the file
            source = os.path.join(dir_of_cert,new_name_certificate + '.png')
            dest = os.path.join(folder_dest,new_name_certificate)
            shutil.copy(source,dest)
            file_name = os.path.join(dest,new_name_certificate)
            os.rename(file_name+'.png',file_name+f'{topic_number}.png')
            print(source)
            print(dest)
            print('sucessfully copied!')            
            # if topic is in certificate, true
            result = True
        
    print(result)   
 
# function gets array with name and all bool values              
def get_row(row_number):
    row = ws[row_number]
    array = []
    for cell in row:
        array.append(cell.value)
        
    return array


max_rows = 27

for row_number in range(1,max_rows):
    # array of names and bool
    array = get_row(row_number)
    # first element is name
    name_of_participant = array[0]

    for i in range(1,9):
        cert_finder(name_of_participant,i)
    # for x in range(1,11):
        # index of 1-4 is day1
        # index of 5-7 is day 2
        # index of 8-9 is day 3
        # a function that accepts index number/topic number
        # if true match topic number to respected file location
        # if false, move on!
        # 
        


